import scrapy
from openpyxl import load_workbook,Workbook
from scrapy.linkextractors import LinkExtractor
from scrapy.spiders import CrawlSpider, Rule
from scrapy_redis.spiders import RedisSpider
import json
import copy
import time
from ..items import TrendsGoogleItem
import xlwt
from urllib import parse
import os
class TrendSpiderSpider(RedisSpider):

    name = 'Trend_spider'
    allowed_domains = ['google.com']
    # start_urls = ['http://trends.google.com/']
    redis_key = 'explore_url_ok'
    # rules = (
    #     Rule(LinkExtractor(allow=r'.*'), callback='parse_item', follow=True),
    # )

    def parse(self, response):
        # if response.status!=200:
        #     print(response.url+"状态码不是200，重新访问")
        #     yield scrapy.Request(url=response.url, callback=self.parse)
        content=response.text.replace(")]}'","")
        # print(content)
        content_json=json.loads(content)
        token=content_json['widgets'][0]['token']
        req=content_json['widgets'][0]['request']
        items=TrendsGoogleItem()

        items['keyword']=req['comparisonItem'][0]['complexKeywordsRestriction']['keyword'][0]['value']
        url='https://trends.google.com/trends/api/widgetdata/multiline?hl=zh-CN&tz=-480&req={}&token={}&tz=-480'.format(req, token)
        yield scrapy.Request(url, meta={"items":copy.deepcopy(items)},callback=self.filter,dont_filter=True)


    def filter(self,response):

        items= response.meta['items']
        if response.status!=200:
            print(response.url+"状态码不是200，重新访问")
            yield scrapy.Request(url=response.url, callback=self.filter,meta={"items":items})
        content=response.text.replace(")]}',","")

        content_json=json.loads(content)
        date=content_json['default']['timelineData']
        timestamp = []
        value = []
        for data_e in date:

            timestampstr = int(data_e['time'])
            timestampstr = TimeTrans(timestampstr)
            valuenum = data_e['value'][0]
            # print(timestamp, value, keyword)
            timestamp.append(timestampstr)
            value.append(valuenum)
        items['time']=timestamp
        items['value']=value
        yield items
        # self.data_trend = {"time": timestamp, keyword: value}
        # # print(data_trend)
        # self.filename = "data/"+keyword + '.xlsx'
        #
        # if not os.path.exists(self.filename):
        #     self.i[keyword]=1
        #     sb = Workbook()
        #     sh = sb.create_sheet('sheet1')
        #     sb.save(self.filename)
            # self.i[keyword]+=self.writeDataToExcleFile(data_trend, self.i[keyword],filename,keyword)
        # self.writeDataToExcleFile()
    # def writeDataToExcleFile(self):
    #   # wb=Workbook()
    #   # sheet = wb.active()
    #
    #   wk = load_workbook(self.filename)
    #   wk_name = wk.sheetnames
    #   wk_sheet = wk[wk_name[0]]
    #
    #   # ws=wb['Sheet1']
    #   # sheet.title = 'sheet1'
    #   key1name = list(self.data_trend.keys())[0]
    #   key2name = list(self.data_trend.keys())[1]
    #
    #   for k in range(len(self.data_trend[key1name])):
    #
    #      wk_sheet.cell(row=1,column=1,value="关键词名称")
    #      wk_sheet.cell(row=self.i[self.keyword] + k, column=1, value=self.keyword)
    #      wk_sheet.cell(row=1,column=2,value="时间")
    #      wk_sheet.cell(row=self.i[self.keyword]+k, column=2, value=self.data_trend[key1name][k])
    #      wk_sheet.cell(row=1,column=3,value="频率")
    #      wk_sheet.cell(row=self.i[self.keyword]+k,column=3,value=self.data_trend[key2name][k])
    #   wk.save(self.filename)
    #   self.i[self.keyword]+=len(self.data_trend[key1name])
    #   print("已保存半年，文件名：{}，行号：{}".format(self.filename,self.i[self.keyword]))
      # return len(data_trend[key1name])
def TimeTrans(timenum):
    timeStamp = timenum
    timeArray = time.localtime(timeStamp)
    otherStyleTime = time.strftime("%Y/%m/%d", timeArray)
    return otherStyleTime
